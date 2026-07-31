import os
import openpyxl
from utils.logger import logger
from openpyxl.utils.exceptions import InvalidFileException

CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(CURRENT_DIR)
TEST_DATA_DIR = os.path.join(PROJECT_DIR, "test_data")

def load_workbook_file(file_path):
    if not os.path.exists(file_path):
        logger.error(f"Excel file not found at: {file_path}")
        raise FileNotFoundError(f"Excel file not found at: {file_path}")

    if os.path.getsize(file_path) == 0:
        logger.error(f"Excel workbook is empty at path: {file_path}")
        raise ValueError("Excel workbook is empty")

    try:
        wb = openpyxl.load_workbook(file_path)
        return wb
    except InvalidFileException:
        logger.error(f"Corrupted or invalid Excel file format at path: {file_path}")
        raise ValueError("File is corrupted or not a valid .xlsx file")

def read_sheet_data(file_name, sheet_name):
    file_path = os.path.join(TEST_DATA_DIR, file_name)
    wb = load_workbook_file(file_path)

    # Check if sheet exists
    if sheet_name not in wb.sheetnames:
        logger.error(f"Sheet '{sheet_name}' not found in workbook '{file_name}'. Available sheets: {wb.sheetnames}")
        raise ValueError(f"Sheet '{sheet_name}' not found in workbook. Available sheets: {wb.sheetnames}")

    sheet = wb[sheet_name]

    # Check if sheet contains rows
    if sheet.max_row < 1:
        logger.error(f"Sheet '{sheet_name}' is completely empty.")
        raise ValueError(f"Sheet '{sheet_name}' is empty.")

    # Read header row (Row 1)
    headers = [cell.value for cell in sheet[1] if cell.value is not None]
    
    if not headers:
        logger.error(f"No headers found on Row 1 in sheet '{sheet_name}'.")
        raise ValueError(f"Missing headers in sheet '{sheet_name}'.")

    data_list = []
    # Read remaining data rows starting from Row 2
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Skip completely blank rows
        if not any(row):
            continue
        
        row_dict = {}
        for header, cell_value in zip(headers, row):
            header_clean = str(header).strip() if header else ""
            
            # Format value: strip strings, keep numbers/booleans as string representation
            if cell_value is None:
                val_clean = ""
            else:
                val_clean = str(cell_value).strip()
                
            row_dict[header_clean] = val_clean

        data_list.append(row_dict)

    logger.info(f"Successfully loaded {len(data_list)} rows from Excel sheet '{sheet_name}' in '{file_name}'.")
    return data_list


def write_test_result(file_name, sheet_name, row_idx, status, status_code, exec_time, remarks=""):

    file_path = os.path.join(TEST_DATA_DIR, file_name)
    wb = load_workbook_file(file_path)

    if sheet_name not in wb.sheetnames:
        logger.error(f"Cannot write results. Sheet '{sheet_name}' not found in '{file_name}'.")
        raise ValueError(f"Sheet '{sheet_name}' not found.")

    sheet = wb[sheet_name]

    # Ensure result headers exist on Row 1
    result_headers = ["Result", "StatusCode", "ExecutionTime", "Remarks"]
    
    # Read existing headers on Row 1
    existing_headers = [sheet.cell(row=1, column=col).value for col in range(1, sheet.max_column + 1)]
    
    header_col_map = {}
    for h in result_headers:
        if h in existing_headers:
            header_col_map[h] = existing_headers.index(h) + 1
        else:
            # Append new header column if missing
            new_col_idx = len(existing_headers) + 1
            sheet.cell(row=1, column=new_col_idx, value=h)
            existing_headers.append(h)
            header_col_map[h] = new_col_idx

    # Write values to target row
    sheet.cell(row=row_idx, column=header_col_map["Result"], value=status)
    sheet.cell(row=row_idx, column=header_col_map["StatusCode"], value=status_code)
    sheet.cell(row=row_idx, column=header_col_map["ExecutionTime"], value=exec_time)
    sheet.cell(row=row_idx, column=header_col_map["Remarks"], value=remarks)

    wb.save(file_path)
    logger.info(f"Updated test results in Excel row {row_idx} (Sheet: '{sheet_name}') -> Status: {status}")